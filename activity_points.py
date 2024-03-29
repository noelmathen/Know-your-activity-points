from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
import time
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from bs4 import BeautifulSoup

# Initialize webdriver
URL = "https://www.rajagiritech.ac.in/stud/ktu/Student/"
driver = webdriver.Chrome()
driver.get(URL)

# Login to the website
driver.find_element(By.NAME, "Userid").send_keys('U2109053')
driver.find_element(By.NAME, "Password").send_keys('210825')
driver.find_element(By.XPATH, "//input[@type='submit']").click()
driver.find_element(By.LINK_TEXT, "Activity Point Form").click()

# Get class codes
classCodeDropdown = Select(driver.find_element(By.NAME, "Class_Code"))
classCodes = [option.get_attribute("value") for option in classCodeDropdown.options]

# Create DataFrame for activity points
activityPoints = pd.DataFrame(classCodes, columns=['SemCode'])
activityPoints['Semester'] = "Semester " + activityPoints['SemCode'].str[-3]
activityPoints['Points'] = [0] * len(activityPoints)
activityPoints = activityPoints.sort_values(by='SemCode')

combined_df = []

# Define Excel file name
excel_file = 'activity_points.xlsx'

# Write data to Excel file
with pd.ExcelWriter(excel_file, engine='xlsxwriter') as writer:
    for index, classCode in enumerate(classCodes):
        combined_df.clear()

        # Select class code
        classCodeDropdown = Select(driver.find_element(By.NAME, "Class_Code"))
        classCodeDropdown.select_by_value(classCode)

        # Get activities
        activitiesDropdown = Select(driver.find_element(By.NAME, "ACode"))
        activities = [option.get_attribute("value") for option in activitiesDropdown.options]
        for activity in activities:
            activitiesDropdown = Select(driver.find_element(By.NAME, "ACode"))
            activitiesDropdown.select_by_value(activity)

            # Add activity
            driver.find_element(By.XPATH, "//input[@value='Add Activity']").click()
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.NAME, "B1")))

            try:
                # Extract HTML table
                table = driver.find_element(By.XPATH, "//table[@class='table table-striped']")
            except:
                continue

            # Parse HTML table
            table_html = table.get_attribute("outerHTML")
            soup = BeautifulSoup(table_html, "html.parser")
            df = pd.read_html(str(soup), header=0)[0].iloc[:-1]
            df = df.dropna(how='all').reset_index(drop=True)

            # Extract certificate links
            certificate_links = [f"https://www.rajagiritech.ac.in/stud/ktu/Student/{a['href']}" for a in soup.find_all('a', href=True)]
            df['Certificate (File Size<500kb)'] = certificate_links
            df['Rating By Faculty'] = df['Rating By Faculty'].replace(np.nan, 0)
            df.index += 1
            activityPoints.loc[index, 'Points'] += df['Rating By Faculty'].sum()

            combined_df.append(df)

        # Write data to Excel sheet
        start_row = 0
        sheet_name = activityPoints.loc[index, 'Semester']
        for df in combined_df:
            df.to_excel(writer, sheet_name=sheet_name, startrow=start_row, index=False)
            start_row += len(df) + 2

    # Write activity points to Statistics sheet
    activityPoints = activityPoints.sort_values(by='SemCode').reset_index(drop=True)
    activityPoints.drop('SemCode', axis=1, inplace=True)
    start_row = 0
    activityPoints.to_excel(writer, sheet_name='Statistics', startrow=start_row, index=False)

    # Write total points to Statistics sheet
    total = activityPoints['Points'].sum()
    result = f"{total} "
    df = pd.DataFrame({'Total': [result]})
    start_row += len(activityPoints) + 3
    df.to_excel(writer, sheet_name='Statistics', index=False, startrow=start_row)

time.sleep(2)

# Apply left alignment to all cells in all sheets
wb = load_workbook(filename=excel_file)
for sheetname in wb.sheetnames:
    ws = wb[sheetname]
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='left')

wb.save(excel_file)
